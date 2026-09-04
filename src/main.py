# -*- coding: utf-8 -*-
import os
import sys
import math
import traceback
from pathlib import Path
from datetime import datetime, date, time

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import xlwt


APP_NAME = "XLSX_to_XLS V6.0"
OUT_DIR_NAME = "XLS_轉換完成"
MAX_ROWS = 65536
MAX_COLS = 256
MAX_STYLE_CACHE = 3500


def safe_text(v):
    if v is None:
        return ""
    return str(v)


def unique_path(folder: Path, stem: str) -> Path:
    p = folder / f"{stem}.xls"
    if not p.exists():
        return p
    for i in range(1, 10000):
        p = folder / f"{stem}_converted_{i}.xls"
        if not p.exists():
            return p
    return folder / f"{stem}_converted.xls"


def rgb_to_xlwt_colour(rgb):
    # xlwt supports a fixed palette. Keep mapping deliberately small/stable.
    if not rgb:
        return None
    rgb = rgb[-6:].upper()
    palette = {
        "000000": 0x08, "FFFFFF": 0x09,
        "FF0000": 0x0A, "00FF00": 0x0B, "0000FF": 0x0C,
        "FFFF00": 0x0D, "FF00FF": 0x0E, "00FFFF": 0x0F,
        "800000": 0x10, "008000": 0x11, "000080": 0x12,
        "808000": 0x13, "800080": 0x14, "008080": 0x15,
        "C0C0C0": 0x16, "808080": 0x17,
    }
    return palette.get(rgb)


def style_key(cell):
    f = cell.font
    fill = cell.fill
    a = cell.alignment
    b = cell.border
    n = cell.number_format or "General"

    def side_key(s):
        return (s.style or "", getattr(s.color, "rgb", None) if s.color else None)

    return (
        f.name, round(float(f.sz or 11), 2), bool(f.bold), bool(f.italic),
        f.underline or "", bool(f.strike), getattr(f.color, "rgb", None) if f.color else None,
        fill.fill_type or "", getattr(fill.fgColor, "rgb", None), getattr(fill.bgColor, "rgb", None),
        a.horizontal or "", a.vertical or "", bool(a.wrap_text), int(a.text_rotation or 0),
        side_key(b.left), side_key(b.right), side_key(b.top), side_key(b.bottom),
        n,
    )


def build_xlwt_style(cell, style_cache):
    key = style_key(cell)
    if key in style_cache:
        return style_cache[key]

    # Hard limit to avoid the BIFF8/XF explosion that caused previous versions to fail.
    if len(style_cache) >= MAX_STYLE_CACHE:
        return xlwt.Style.default_style

    st = xlwt.XFStyle()

    f = cell.font
    xf = xlwt.Font()
    xf.name = f.name or "Arial"
    xf.height = int(float(f.sz or 10) * 20)
    xf.bold = bool(f.bold)
    xf.italic = bool(f.italic)
    xf.struck_out = bool(f.strike)
    if f.underline:
        xf.underline = xlwt.Font.UNDERLINE_SINGLE
    c = getattr(f.color, "rgb", None) if f.color else None
    ci = rgb_to_xlwt_colour(c)
    if ci is not None:
        xf.colour_index = ci
    st.font = xf

    a = cell.alignment
    xa = xlwt.Alignment()
    hmap = {
        "left": xlwt.Alignment.HORZ_LEFT,
        "center": xlwt.Alignment.HORZ_CENTER,
        "right": xlwt.Alignment.HORZ_RIGHT,
        "justify": xlwt.Alignment.HORZ_JUSTIFIED,
    }
    vmap = {
        "top": xlwt.Alignment.VERT_TOP,
        "center": xlwt.Alignment.VERT_CENTER,
        "bottom": xlwt.Alignment.VERT_BOTTOM,
    }
    if a.horizontal in hmap:
        xa.horz = hmap[a.horizontal]
    if a.vertical in vmap:
        xa.vert = vmap[a.vertical]
    xa.wrap = 1 if a.wrap_text else 0
    if a.text_rotation:
        xa.rota = int(a.text_rotation)
    st.alignment = xa

    fill = cell.fill
    xp = xlwt.Pattern()
    if fill.fill_type and getattr(fill.fgColor, "rgb", None):
        fi = rgb_to_xlwt_colour(fill.fgColor.rgb)
        if fi is not None:
            xp.pattern = xlwt.Pattern.SOLID_PATTERN
            xp.pattern_fore_colour = fi
    st.pattern = xp

    border = cell.border
    xb = xlwt.Borders()
    line_map = {
        "thin": xlwt.Borders.THIN,
        "medium": xlwt.Borders.MEDIUM,
        "dashed": xlwt.Borders.DASHED,
        "dotted": xlwt.Borders.DOTTED,
        "thick": xlwt.Borders.THICK,
        "double": xlwt.Borders.DOUBLE,
        "hair": xlwt.Borders.HAIR,
    }
    for src, attr in [
        (border.left, "left"), (border.right, "right"),
        (border.top, "top"), (border.bottom, "bottom")
    ]:
        if src.style in line_map:
            setattr(xb, attr, line_map[src.style])
    st.borders = xb

    try:
        st.num_format_str = cell.number_format or "General"
    except Exception:
        st.num_format_str = "General"

    style_cache[key] = st
    return st


def write_value(ws_out, row, col, cell, style):
    v = cell.value
    if v is None:
        ws_out.write(row, col, "", style)
        return

    if isinstance(v, bool):
        ws_out.write(row, col, bool(v), style)
    elif isinstance(v, (int, float)) and not isinstance(v, bool):
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            ws_out.write(row, col, safe_text(v), style)
        else:
            ws_out.write(row, col, v, style)
    elif isinstance(v, (datetime, date, time)):
        ws_out.write(row, col, v, style)
    elif isinstance(v, str) and v.startswith("="):
        try:
            ws_out.write(row, col, xlwt.Formula(v[1:]), style)
        except Exception:
            # Preserve formula text if BIFF8 formula parser cannot understand it.
            ws_out.write(row, col, v, style)
    else:
        ws_out.write(row, col, safe_text(v), style)


def convert_one(src_path: str) -> Path:
    src = Path(src_path)
    wb_in = load_workbook(
        src,
        data_only=False,
        read_only=False,
        keep_links=False,
        keep_vba=False
    )

    wb_out = xlwt.Workbook(style_compression=2)
    style_cache = {}

    for ws_in in wb_in.worksheets:
        max_row = ws_in.max_row or 1
        max_col = ws_in.max_column or 1

        if max_row > MAX_ROWS:
            raise ValueError(f"{ws_in.title} 超過 XLS 65536 列限制：{max_row}")
        if max_col > MAX_COLS:
            raise ValueError(f"{ws_in.title} 超過 XLS 256 欄限制：{max_col}")

        title = (ws_in.title or "Sheet")[:31]
        ws_out = wb_out.add_sheet(title, cell_overwrite_ok=True)

        # Widths
        for col_idx in range(1, max_col + 1):
            letter = get_column_letter(col_idx)
            dim = ws_in.column_dimensions.get(letter)
            if dim and dim.width:
                width = max(1, min(255, float(dim.width)))
                ws_out.col(col_idx - 1).width = int(width * 256)

        # Heights
        for row_idx, dim in ws_in.row_dimensions.items():
            if dim.height and 1 <= row_idx <= MAX_ROWS:
                ws_out.row(row_idx - 1).height_mismatch = True
                ws_out.row(row_idx - 1).height = int(float(dim.height) * 20)

        merged_top_left = {}
        merged_interior = set()
        for rng in ws_in.merged_cells.ranges:
            min_col, min_row, max_col_m, max_row_m = rng.bounds
            if max_row_m > MAX_ROWS or max_col_m > MAX_COLS:
                continue
            merged_top_left[(min_row, min_col)] = (max_row_m, max_col_m)
            for rr in range(min_row, max_row_m + 1):
                for cc in range(min_col, max_col_m + 1):
                    if (rr, cc) != (min_row, min_col):
                        merged_interior.add((rr, cc))

        for row in ws_in.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                if (cell.row, cell.column) in merged_interior:
                    continue

                style = build_xlwt_style(cell, style_cache)

                if (cell.row, cell.column) in merged_top_left:
                    max_row_m, max_col_m = merged_top_left[(cell.row, cell.column)]
                    value = cell.value
                    if isinstance(value, str) and value.startswith("="):
                        try:
                            value = xlwt.Formula(value[1:])
                        except Exception:
                            pass
                    ws_out.write_merge(
                        cell.row - 1, max_row_m - 1,
                        cell.column - 1, max_col_m - 1,
                        value if value is not None else "",
                        style
                    )
                else:
                    write_value(ws_out, cell.row - 1, cell.column - 1, cell, style)

    out_dir = src.parent / OUT_DIR_NAME
    out_dir.mkdir(exist_ok=True)
    dst = unique_path(out_dir, src.stem)
    wb_out.save(str(dst))

    # Validate true OLE/BIFF8 container signature.
    with open(dst, "rb") as f:
        sig = f.read(8)
    if sig != bytes.fromhex("D0CF11E0A1B11AE1"):
        try:
            dst.unlink()
        except Exception:
            pass
        raise ValueError("輸出驗證失敗：不是有效 OLE/BIFF8 XLS")

    return dst


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_NAME + "｜產線簡易穩定版")
        self.geometry("720x430")
        self.minsize(650, 390)
        self.files = []

        self.columnconfigure(0, weight=1)
        self.rowconfigure(2, weight=1)

        head = ttk.Frame(self, padding=14)
        head.grid(row=0, column=0, sticky="ew")
        head.columnconfigure(0, weight=1)

        ttk.Label(
            head,
            text="XLSX → 真正 Excel 97-2003 XLS",
            font=("Microsoft JhengHei UI", 16, "bold")
        ).grid(row=0, column=0, sticky="w")

        ttk.Label(
            head,
            text="V6.0｜產線資料表優先｜不需要 Excel / LibreOffice / COM",
            font=("Microsoft JhengHei UI", 10)
        ).grid(row=1, column=0, sticky="w", pady=(4, 0))

        buttons = ttk.Frame(self, padding=(14, 0, 14, 8))
        buttons.grid(row=1, column=0, sticky="ew")

        ttk.Button(buttons, text="選擇 XLSX（可多選）", command=self.choose).pack(side="left")
        ttk.Button(buttons, text="清除清單", command=self.clear).pack(side="left", padx=8)
        self.btn_convert = ttk.Button(buttons, text="開始轉換", command=self.convert)
        self.btn_convert.pack(side="right")

        body = ttk.Frame(self, padding=(14, 0, 14, 10))
        body.grid(row=2, column=0, sticky="nsew")
        body.columnconfigure(0, weight=1)
        body.rowconfigure(0, weight=1)

        self.listbox = tk.Listbox(body, font=("Microsoft JhengHei UI", 10))
        self.listbox.grid(row=0, column=0, sticky="nsew")
        sb = ttk.Scrollbar(body, orient="vertical", command=self.listbox.yview)
        sb.grid(row=0, column=1, sticky="ns")
        self.listbox.configure(yscrollcommand=sb.set)

        foot = ttk.Frame(self, padding=14)
        foot.grid(row=3, column=0, sticky="ew")
        foot.columnconfigure(0, weight=1)
        self.status = tk.StringVar(value="請先選擇要轉換的 XLSX 檔案。")
        ttk.Label(foot, textvariable=self.status).grid(row=0, column=0, sticky="w")
        ttk.Label(
            foot,
            text="輸出：來源檔案旁的「XLS_轉換完成」資料夾",
            font=("Microsoft JhengHei UI", 9)
        ).grid(row=1, column=0, sticky="w", pady=(5, 0))

    def choose(self):
        files = filedialog.askopenfilenames(
            title="選擇 XLSX（可多選）",
            filetypes=[("Excel XLSX", "*.xlsx")]
        )
        if not files:
            return
        self.files = list(files)
        self.listbox.delete(0, tk.END)
        for p in self.files:
            self.listbox.insert(tk.END, p)
        self.status.set(f"已選擇 {len(self.files)} 個檔案。")

    def clear(self):
        self.files = []
        self.listbox.delete(0, tk.END)
        self.status.set("已清除。")

    def convert(self):
        if not self.files:
            messagebox.showwarning(APP_NAME, "請先選擇 XLSX 檔案。")
            return

        self.btn_convert.configure(state="disabled")
        self.update_idletasks()

        ok = 0
        errors = []
        for i, p in enumerate(self.files, 1):
            self.status.set(f"正在轉換 {i}/{len(self.files)}：{Path(p).name}")
            self.update()
            try:
                convert_one(p)
                ok += 1
            except Exception as e:
                errors.append(f"{Path(p).name}\n{e}")

        self.btn_convert.configure(state="normal")
        fail = len(errors)
        self.status.set(f"完成：成功 {ok}，失敗 {fail}")

        if errors:
            messagebox.showerror(
                APP_NAME,
                f"轉換完成\n\n成功：{ok}\n失敗：{fail}\n\n" +
                "\n\n".join(errors[:10])
            )
        else:
            messagebox.showinfo(
                APP_NAME,
                f"轉換完成\n\n成功：{ok}\n失敗：0\n\n"
                f"輸出位置：來源檔案旁的「{OUT_DIR_NAME}」"
            )


if __name__ == "__main__":
    try:
        App().mainloop()
    except Exception:
        messagebox.showerror(APP_NAME, traceback.format_exc())
